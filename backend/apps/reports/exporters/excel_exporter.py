"""
ExcelExporter — Strategy for .xlsx via openpyxl (lazy import). SOLID: SRP·LSP·OCP.
openpyxl is imported inside export() so the module loads even if the lib is absent;
a clear error is raised only when an Excel export is actually requested.
"""

from __future__ import annotations

from apps.reports.exporters.spreadsheet_cell import spreadsheet_safe_value
from apps.reports.interfaces import IReportExporter


class ExcelExporter(IReportExporter):

    def export(self, rows: list[dict], columns: list[str]) -> bytes:
        try:
            from openpyxl import Workbook  # noqa: PLC0415
            from openpyxl.styles import Font  # noqa: PLC0415
            from openpyxl.utils import get_column_letter  # noqa: PLC0415
        except ImportError as exc:
            raise RuntimeError("Instala openpyxl para exportar a Excel.") from exc

        import io  # noqa: PLC0415
        wb = Workbook()
        ws = wb.active
        ws.title = "Tickets"
        ws.append(columns)
        for row in rows:
            ws.append([spreadsheet_safe_value(row.get(c, "")) for c in columns])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for column_index, column in enumerate(columns, start=1):
            values = [column, *(row.get(column, "") for row in rows)]
            content_width = max(len(str(value or "")) for value in values) + 2
            letter = get_column_letter(column_index)
            ws.column_dimensions[letter].width = min(max(content_width, 8), 54)
            if column in {"numero", "ruc"}:
                for cell in ws.iter_cols(
                    min_col=column_index,
                    max_col=column_index,
                    min_row=2,
                ):
                    cell[0].number_format = "@"
                    cell[0].quotePrefix = True
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def get_extension(self) -> str:
        return "xlsx"

    def get_mime_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
