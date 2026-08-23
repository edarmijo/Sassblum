"""
Tests for report exporters + ExporterFactory (no database required).
Run: pytest apps/reports/tests/test_exporters.py -v
"""

import csv
from io import BytesIO
from io import StringIO

import pytest
from openpyxl import load_workbook

from apps.reports.exporters import CSVExporter, ExcelExporter, PDFExporter
from core.factories.exporter_factory import ExporterFactory

ROWS = [
    {
        "numero": "T-2026-0001",
        "creado_en": "2026-08-23 10:30",
        "usuario": "Victoria Pinto",
        "empresa": "SassBlum",
        "ruc": "0999999999001",
        "asunto": "Revisión mensual",
        "estado": "Nuevo",
        "prioridad": "Alta",
        "servicio": "Soporte",
        "asignado": "tecnico@sassblum.com",
    },
    {
        "numero": "T-2026-0002",
        "creado_en": "2026-08-23 11:45",
        "usuario": "Patricio Blum",
        "empresa": "DULCENAC S.A.",
        "ruc": "0992338547001",
        "asunto": "Mantenimiento preventivo",
        "estado": "Cerrado",
        "prioridad": "Baja",
        "servicio": "Infraestructura",
        "asignado": "",
    },
]
COLUMNS = [
    "numero", "creado_en", "usuario", "empresa", "ruc",
    "asunto", "estado", "prioridad", "servicio", "asignado",
]


class TestExcelExporter:
    def test_extension_and_mime(self):
        exporter = ExcelExporter()
        assert exporter.get_extension() == "xlsx"
        assert exporter.get_mime_type() == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    def test_export_includes_columns_and_rows(self):
        workbook = load_workbook(BytesIO(ExcelExporter().export(ROWS, COLUMNS)))
        values = list(workbook.active.values)

        assert values[0] == tuple(COLUMNS)
        assert values[1] == tuple(ROWS[0][column] for column in COLUMNS)
        assert values[2][:-1] == tuple(ROWS[1][column] for column in COLUMNS[:-1])
        assert values[2][-1] is None
        assert workbook.active["A1"].font.bold is True
        assert workbook.active["A2"].number_format == "@"
        assert workbook.active["E2"].number_format == "@"
        assert workbook.active["E2"].quotePrefix is True
        assert workbook.active.column_dimensions["F"].width >= len(ROWS[1]["asunto"])
        assert workbook.active.freeze_panes == "A2"
        assert workbook.active.auto_filter.ref == "A1:J3"

    def test_export_keeps_formula_like_text_literal(self):
        rows = [{"asunto": "=HYPERLINK(\"https://example.com\")"}]
        workbook = load_workbook(BytesIO(ExcelExporter().export(rows, ["asunto"])))

        assert workbook.active["A2"].value.startswith("'=")


class TestCSVExporter:
    def test_extension_mime_columns_and_utf8_rows(self):
        exporter = CSVExporter()
        content = exporter.export(ROWS, COLUMNS).decode("utf-8-sig")
        values = list(csv.reader(StringIO(content)))

        assert exporter.get_extension() == "csv"
        assert exporter.get_mime_type() == "text/csv; charset=utf-8"
        assert values[0] == COLUMNS
        assert values[1] == [str(ROWS[0][column]) for column in COLUMNS]
        assert values[2] == [str(ROWS[1][column]) for column in COLUMNS]

    def test_export_quotes_cells_and_keeps_formula_like_text_literal(self):
        rows = [{"asunto": "=1+1, con coma"}]
        values = list(csv.reader(StringIO(
            CSVExporter().export(rows, ["asunto"]).decode("utf-8-sig")
        )))

        assert values == [["asunto"], ["'=1+1, con coma"]]


class TestExporterFactory:
    def test_build_excel(self):
        assert isinstance(ExporterFactory.build("excel"), ExcelExporter)

    def test_build_is_case_insensitive(self):
        assert isinstance(ExporterFactory.build("EXCEL"), ExcelExporter)

    def test_build_csv(self):
        assert isinstance(ExporterFactory.build("csv"), CSVExporter)

    def test_unknown_format_raises(self):
        with pytest.raises(ValueError):
            ExporterFactory.build("xml")

    def test_pdf_and_excel_resolve_without_libs(self):
        # The classes resolve even if reportlab/openpyxl aren't installed;
        # the ImportError only surfaces when export() is called.
        assert ExporterFactory.build("pdf") is not None
        assert ExporterFactory.build("excel") is not None


class TestPDFExporter:
    def test_export_uses_landscape_letter_page(self):
        columns = [
            "numero", "creado_en", "usuario", "empresa", "ruc",
            "asunto", "estado", "prioridad", "servicio", "asignado",
        ]
        rows = [{column: f"Contenido extenso para {column} " * 4 for column in columns}]

        content = PDFExporter().export(rows, columns)

        assert content.startswith(b"%PDF-")
        assert b"/MediaBox [ 0 0 792 612 ]" in content

    def test_export_handles_multiple_pages_and_long_cells(self):
        columns = [
            "numero", "creado_en", "usuario", "empresa", "ruc",
            "asunto", "estado", "prioridad", "servicio", "asignado",
        ]
        rows = [
            {column: f"Fila {index}: valor largo de {column} " * 3 for column in columns}
            for index in range(80)
        ]

        content = PDFExporter().export(rows, columns)

        assert content.startswith(b"%PDF-")
        assert len(content) > 1_000
